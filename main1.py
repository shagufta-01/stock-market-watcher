import yfinance as yf
import schedule
import pygame
import time
from openpyxl import Workbook, load_workbook
from plyer import notification
import os
from datetime import datetime
import mplfinance as mpf
import matplotlib.pyplot as plt

TICKER = "RELIANCE.NS"
EXCEL_FILE = "StockMarket_log.xlsx"

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Company", "Ticker", "Open", "Low", "High", "Close", "Matched"])
    wb.save(EXCEL_FILE)

def log_to_excel(company, ticker, open_p, low_p, high_p, close_p, matched):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        company,
        ticker,
        open_p,
        low_p,
        high_p,
        close_p,
        "Yes" if matched else "No"
    ])
    wb.save(EXCEL_FILE)

def notify_user(title, message):
    notification.notify(
        title=title,
        message=message,
        timeout=5
    )

def check_market_condition():
    stock = yf.Ticker(TICKER)
    hist = stock.history(period="1d", interval="1m")
    if hist.empty:
        print("No data available.")
        return

    latest = hist.iloc[-1]
    candle_time = latest.name.strftime('%Y-%m-%d %H:%M:%S')
    open_price = latest['Open']
    low_price = latest['Low']
    high_price = latest['High']
    close_price = latest['Close']
    company_name = stock.info.get("longName", TICKER)

    print(f"\n📈 Checking: {company_name} ({TICKER})")
    print(f"🕒 Candle Time: {candle_time}")
    print("📊 Prices =>", {
        "open": open_price,
        "low": low_price,
        "high": high_price,
        "close": close_price
    })

    matched = open_price == low_price or high_price == close_price

    if matched:
        print("🔔 Condition matched! Playing sound...")
        try:
            pygame.mixer.init()
            pygame.mixer.music.load(os.path.join(os.getcwd(), "pop-1.mp3"))
            pygame.mixer.music.play()
            time.sleep(2)
            pygame.mixer.music.stop()
        except Exception as e:
            print("⚠️ Error playing sound:", e)
        notify_user("Market Signal", f"{company_name} ({TICKER}) matched condition!")

    else:
        print("❌ Condition not matched.")

    log_to_excel(company_name, TICKER, open_price, low_price, high_price, close_price, matched)

# Global figure handle to keep chart window open
candle_fig = None

def update_candlestick_graph():
    global candle_fig
    stock = yf.Ticker(TICKER)
    data = stock.history(period="1d", interval="5m")

    if data.empty:
        print("⚠️ No data for candlestick chart.")
        return

    if candle_fig:
        plt.close(candle_fig)

    candle_fig, _ = mpf.plot(
        data,
        type='candle',
        style='yahoo',
        title=f"{TICKER} - Live Intraday Candlestick",
        volume=True,
        returnfig=True,
        block=False
    )
    plt.pause(0.1)  # Allow matplotlib to render the updated chart

# Schedule every 5 minutes
schedule.every(5).minutes.do(check_market_condition)
schedule.every(5).minutes.do(update_candlestick_graph)

print("🚀 Live Market Watcher with Candlestick Chart Started...")

check_market_condition()
update_candlestick_graph()

while True:
    schedule.run_pending()
    time.sleep(1)
