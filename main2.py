import yfinance as yf
import pygame
import time
from openpyxl import Workbook, load_workbook

import os
from datetime import datetime
import mplfinance as mpf
import matplotlib.pyplot as plt
from win10toast import ToastNotifier

# === Config ===
# TICKER = "RELIANCE.NS"
# TICKER = "SBIN.NS"
# TICKER = "BAJFINANCE.NS"
TICKER="BAJFINANCE.NS"
# TICKER="MUTHOOTFIN.NS"
# TICKER="KOTAKBANK.NS"
EXCEL_FILE = "StockMarket01_log.xlsx"
last_checked_candle_time = None
toaster = ToastNotifier()

# === Excel Setup ===
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Company", "Ticker", "Open", "Low", "High", "Close", "Matched"])
    wb.save(EXCEL_FILE)

def log_to_excel(company, ticker, open_p, low_p, high_p, close_p, matched):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        company,
        ticker,
        open_p,
        low_p,
        high_p,
        close_p,
        "Yes" if matched else "No"
    ])
    wb.save(EXCEL_FILE)

def notify_user(title, message):
    toaster.show_toast(
        title,
        message,
        duration=5,
        threaded=True
    )
    while toaster.notification_active():  # Ensures it gets time to show
        time.sleep(0.1)

def is_equal(a, b, tol=0.01):
    return abs(a - b) < tol

def check_new_5min_candle():
    global last_checked_candle_time

    stock = yf.Ticker(TICKER)
    hist = stock.history(period="1d", interval="5m")

    if len(hist) < 6:
        print("⛔ Not enough candles to analyze.")
        return

    latest_candle = hist.iloc[-1]
    latest_time = latest_candle.name.strftime('%Y-%m-%d %H:%M:%S')

    if latest_time == last_checked_candle_time:
        # print(f"⏳ No new candle yet: {latest_time}")
        return

    last_checked_candle_time = latest_time

    open_p = latest_candle['Open']
    low_p = latest_candle['Low']
    high_p = latest_candle['High']
    close_p = latest_candle['Close']
    company_name = stock.info.get("longName", TICKER)

    # Print OHLC values
    print(f"\n🕒 New 5-min Candle: {latest_time}")
    print(f"📊 OHLC - Open: {open_p}, Low: {low_p}, High: {high_p}, Close: {close_p}")

    open_eq_low = is_equal(open_p, low_p)
    high_eq_close = is_equal(high_p, close_p)

    last_5 = hist.iloc[-6:-1]  # last 5 complete candles
    bearish_5 = all(row['Open'] > row['Close'] for _, row in last_5.iterrows())

    print(f"🔍 open==low: {open_eq_low}, high==close: {high_eq_close}, 5 bearish: {bearish_5}")

    matched = open_eq_low and high_eq_close and bearish_5

    if matched:
        print("✅ All 3 conditions matched!")

        # Play sound
        try:
            pygame.mixer.init()
            pygame.mixer.music.load(os.path.join(os.getcwd(), "pop-1.mp3"))
            pygame.mixer.music.play()
            time.sleep(2)
            pygame.mixer.music.stop()
        except Exception as e:
            print("⚠️ Error playing sound:", e)

        # Show notification
        notify_user("Market Signal", f"{company_name} matched at {latest_time}")

        # Log to Excel
        log_to_excel(company_name, TICKER, open_p, low_p, high_p, close_p, True)

        # Show candlestick chart
        try:
            plot_hist = hist[-10:]  # last 10 candles
            mpf.plot(plot_hist, type='candle', style='charles',
                     title=f"{company_name} - Last 10 Candles",
                     ylabel='Price',
                     block=False)
        except Exception as e:
            print("⚠️ Error displaying chart:", e)

    else:
        print("❌ Conditions not matched.")
        log_to_excel(company_name, TICKER, open_p, low_p, high_p, close_p, False)

def run_forever():
    print("🚀 Running Live Candle Watcher...")
    while True:
        check_new_5min_candle()
        time.sleep(10)  # check every 10 seconds

run_forever()
