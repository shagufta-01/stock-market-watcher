import yfinance as yf
import schedule
import time
from playsound import playsound
from openpyxl import Workbook, load_workbook
from plyer import notification
import os
from datetime import datetime

# STOCK TICKER
TICKER = "RELIANCE.NS"

# Excel File
EXCEL_FILE = "market_log.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Open", "Low", "High", "Close", "Matched"])
    wb.save(EXCEL_FILE)

def log_to_excel(open_p, low_p, high_p, close_p, matched):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        open_p,
        low_p,
        high_p,
        close_p,
        "Yes" if matched else "No"
    ])
    wb.save(EXCEL_FILE)

def notify_user(title, message):
    notification.notify(
        title=title,
        message=message,
        timeout=5
    )

def check_market_condition():
    stock = yf.Ticker(TICKER)
    hist = stock.history(period="1d", interval="1m")
    if hist.empty:
        print("No data available.")
        return

    latest = hist.iloc[-1]
    candle_time = latest.name.strftime('%Y-%m-%d %H:%M:%S')
    open_price = latest['Open']
    low_price = latest['Low']
    high_price = latest['High']
    close_price = latest['Close']

    company_name = stock.info.get("longName", TICKER)
    print(f"\n📈 Checking: {company_name} ({TICKER})")
    print(f"🕒 Candle Time: {candle_time}")
    print("📊 Prices =>", {
        "open": open_price,
        "low": low_price,
        "high": high_price,
        "close": close_price
    })

    matched = open_price == low_price and high_price == close_price
    log_to_excel(open_price, low_price, high_price, close_price, matched)

    if matched:
        print("🔔 Condition matched! Playing sound...")
        try:
            playsound(os.path.join(os.getcwd(), "alert.mp3"))
        except Exception as e:
            print("⚠️ Error playing sound:", e)
        notify_user("Market Signal", f"{company_name} matched condition!")
    else:
        print("❌ Condition not matched.")

    stock = yf.Ticker(TICKER)  # ✅ Moved inside function
    hist = stock.history(period="1d", interval="1m")
    if hist.empty:
        print("No data available.")
        return

    latest = hist.iloc[-1]
    open_price = latest['Open']
    low_price = latest['Low']
    high_price = latest['High']
    close_price = latest['Close']

    print("Checking:", {
        "open": open_price,
        "low": low_price,
        "high": high_price,
        "close": close_price
    })

    if open_price == low_price and high_price == close_price:
        print("🔔 Condition matched! Playing sound...")
        playsound('alert.mp3')
        notify_user("Market Signal", f"{TICKER} matched condition!")
        log_to_excel(open_price, low_price, high_price, close_price, True)
    else:
        print("❌ Condition not matched.")
        log_to_excel(open_price, low_price, high_price, close_price, False)

# Schedule every 5 minutes
schedule.every(5).minutes.do(check_market_condition)

print("🚀 Market Watcher with Sound Started...")
while True:
    schedule.run_pending()
    time.sleep(1)
